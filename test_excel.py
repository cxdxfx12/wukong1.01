#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
import sys

def read_excel_headers(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    # 读取表头（第一行）
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value) if cell.value else "")

    print("=== 表头列 ===")
    for i, h in enumerate(headers):
        print(f"列{i+1}: {h}")

    # 读取前5行数据样本
    print("\n=== 前5行数据样本 ===")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=6), start=2):
        print(f"行{row_idx}:")
        for i, cell in enumerate(row):
            if i < len(headers):
                print(f"  {headers[i]}: {cell.value}")

    wb.close()

if __name__ == "__main__":
    filepath = "蜜丝婷-4月发件账单表1.xlsx"
    read_excel_headers(filepath)