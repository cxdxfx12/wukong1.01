import openpyxl
import math
import time

def normalize_province(province):
    if not province:
        return ''
    p = province.strip()
    for suffix in ['省', '市', '自治区', '特别行政区']:
        if p.endswith(suffix):
            p = p[:-len(suffix)]
            break
    return p.strip()

def calculate_freight(weight, province):
    region_mapping = {
        '江苏': 1, '浙江': 1, '安徽': 1, '上海': 1,
        '山东': 2, '广东': 2, '福建': 2, '北京': 2,
        '河南': 2, '湖北': 2, '湖南': 2, '江西': 2,
        '天津': 2, '河北': 2,
        '山西': 3, '广西': 3, '四川': 3, '重庆': 3,
        '陕西': 3, '贵州': 3, '辽宁': 3, '吉林': 3,
        '黑龙江': 3, '云南': 3,
        '海南': 4, '甘肃': 4, '青海': 4, '内蒙古': 4,
        '宁夏': 4,
        '新疆': 5, '西藏': 5
    }

    price_rules = {
        1: [(0, 0.5, 2.26, 0), (0.51, 1, 2.46, 0), (1, 2, 3.56, 0), (2, 3, 4.76, 0), (3, 30, 3.76, 0.8), (30, -1, 3.86, 0.8)],
        2: [(0, 0.5, 2.26, 0), (0.51, 1, 2.46, 0), (1, 2, 3.56, 0), (2, 3, 4.76, 0), (3, 30, 3.76, 1.1), (30, -1, 4.06, 1.3)],
        3: [(0, 0.5, 2.26, 0), (0.51, 1, 2.46, 0), (1, 2, 3.56, 0), (2, 3, 4.76, 0), (3, 30, 3.76, 1.5), (30, -1, 4.06, 1.6)],
        4: [(0, 0.5, 2.56, 0), (0.51, 1, 3.56, 0), (1, 2, 4.06, 0), (2, 3, 5.06, 0), (3, 30, 3.76, 2.5), (30, -1, 4.06, 4.3)],
        5: [(0, 0.5, 10, 0), (0.51, 1, 13, 0), (1, 2, 20, 0), (2, 3, 25, 0), (3, 30, 15, 15), (30, -1, 15, 15)]
    }

    normalized = normalize_province(province)
    region = region_mapping.get(normalized, 2)
    rules = price_rules[region]

    for seg in rules:
        min_w, max_w, first_price, add_price = seg
        if max_w < 0:
            max_w = float('inf')
        if weight >= min_w and weight <= max_w:
            if add_price == 0:
                return first_price
            if max_w == float('inf') and min_w >= 30:
                first_weight = 3.0
            else:
                first_weight = min_w
            extra = weight - first_weight
            extra_kg = math.ceil(extra)
            return first_price + extra_kg * add_price
    return 0

# 测试用例
test_cases = [
    (0.39, '江苏省', 2.26),
    (0.15, '北京', 2.26),
    (0.20, '江西省', 2.26),
    (4.0, '江苏省', 4.56),
    (31.0, '江苏省', 26.26),
    (40.0, '江苏省', 33.46),
    (1.5, '广东省', 3.56),
    (35.0, '新疆维吾尔自治区', 15 + math.ceil(35-3) * 15),
]

print('==== 计算逻辑验证 ====')
all_pass = True
for weight, province, expected in test_cases:
    result = calculate_freight(weight, province)
    status = 'OK' if abs(result - expected) < 0.01 else 'FAIL'
    if status == 'FAIL':
        all_pass = False
    print(f'{status} | {province} {weight}kg -> {result:.2f}元 (预期: {expected:.2f}元)')

result_str = '全部通过' if all_pass else '有失败'
print(f'\n计算验证: {result_str}')

# 文件统计
files = [
    'c:/Users/ccc/Desktop/帐单/珀莱雅-4月发件账单.xlsx',
    'c:/Users/ccc/Desktop/帐单/蜜丝婷-4月发件账单表1.xlsx',
    'c:/Users/ccc/Desktop/帐单/蜜丝婷-4月发件账单表2.xlsx',
    'c:/Users/ccc/Desktop/帐单/蜜丝婷-5月发件账单.xlsx'
]

print('\n==== 端到端测试 ====')
total_rows = 0
total_freight = 0
customers = set()

for f in files:
    fname = f.split('/')[-1]
    print(f'\n处理: {fname}')
    start = time.time()

    wb = openpyxl.load_workbook(f, read_only=True)
    ws = wb.active
    rows = ws.max_row - 1
    total_rows += rows

    # 采样前100条计算
    sample_freight = 0
    sample_count = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=min(101, ws.max_row))):
        weight = row[2].value if row[2].value else 0
        province = row[3].value if row[3].value else ''
        customer = row[9].value if row[9].value else ''
        if customer:
            customers.add(customer)
        if province and weight:
            freight = calculate_freight(weight, province)
            sample_freight += freight
            sample_count += 1

    elapsed = time.time() - start
    print(f'  总行数: {rows}')
    print(f'  采样100条运费合计: {sample_freight:.2f}元')
    print(f'  采样平均运费: {sample_freight/sample_count:.2f}元')
    print(f'  处理耗时: {elapsed:.1f}s')

print(f'\n==== 汇总 ====')
print(f'总记录数: {total_rows}')
print(f'客户列表: {customers}')
avg = sample_freight/sample_count if sample_count > 0 else 0
print(f'预估总运费(按采样均值推算): {avg*total_rows:.0f}元')
